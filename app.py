import os
import cv2
import sqlite3
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file
import numpy as np
from ultralytics import YOLO
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage, ImageDraw, ImageFont

app = Flask(__name__)
model = YOLO('yolov8m.pt') 

IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.webp'}
VIDEO_EXTENSIONS = {'.mp4', '.avi', '.mov', '.mkv'}

def init_db():
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp DATETIME,
            duck_count INTEGER,
            image_path TEXT
        )
    ''')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_timestamp ON requests(timestamp)')
    conn.commit()
    conn.close()

init_db()
os.makedirs('static', exist_ok=True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_images():
    if 'images' not in request.files:
        return jsonify({'error': 'Нет файлов в запросе'}), 400
        
    files = request.files.getlist('images')
    results_data = []
    
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()

    cursor.execute('DELETE FROM requests')
    conn.commit()
    
    try:
        font = ImageFont.truetype("arial.ttf", 16)
    except:
        font = ImageFont.load_default()
    
    for file in files:
        if file.filename == '':
            continue
            
        ext = os.path.splitext(file.filename)[1].lower()
        safe_filename = "".join([c for c in file.filename if c.isalnum() or c in ['.', '_', '-']])
        timestamp_str = datetime.now().strftime('%Y%m%d%H%M%S')
        
        if ext in IMAGE_EXTENSIONS:
            img = cv2.imdecode(np.frombuffer(file.read(), np.uint8), cv2.IMREAD_COLOR)
            results = model(img, classes=[14], conf=0.15, iou=0.45)
            boxes = results[0].boxes
            duck_count = len(boxes)
            
            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            pil_img = PILImage.fromarray(img_rgb)
            draw = ImageDraw.Draw(pil_img)
            
            for idx, box in enumerate(boxes, start=1):
                x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                draw.rectangle([x1, y1, x2, y2], outline=(0, 255, 0), width=3)
                text = f"Утка {idx}"
                try: text_w, text_h = draw.textbbox((0, 0), text, font=font)[2:]
                except: text_w, text_h = 60, 18
                draw.rectangle([x1, y1 - text_h - 4, x1 + text_w + 6, y1], fill=(0, 255, 0))
                draw.text((x1 + 3, y1 - text_h - 2), text, fill=(0, 0, 0), font=font)
                
            output_img = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
            filepath = os.path.join('static', f"res_{timestamp_str}_{safe_filename}")
            cv2.imwrite(filepath, output_img)
            
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute('INSERT INTO requests (timestamp, duck_count, image_path) VALUES (?, ?, ?)', 
                           (timestamp, duck_count, filepath))
            
            results_data.append({
                'filename': file.filename,
                'count': duck_count,
                'image_url': f"/{filepath}",
                'is_video': False
            })
            
        elif ext in VIDEO_EXTENSIONS:
            temp_input_path = os.path.join('static', f"tmp_{timestamp_str}_{safe_filename}")
            file.save(temp_input_path)
            
            cap = cv2.VideoCapture(temp_input_path)
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            fps = cap.get(cv2.CAP_PROP_FPS) or 30.0

            video_out_name = f"res_{timestamp_str}_{os.path.splitext(safe_filename)[0]}.mp4"
            video_out_path = os.path.join('static', video_out_name)
            
            fourcc = cv2.VideoWriter_fourcc(*'mp4v')
            out = cv2.VideoWriter(video_out_path, fourcc, fps, (width, height))
            
            max_duck_count = 0
            best_frame_bgr = None
            
            img_size = 1088 if max(width, height) >= 1920 else 736
            
            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    break
                    
                results = model.track(frame, persist=True, classes=[14], conf=0.15, iou=0.45, imgsz=img_size, verbose=False)
                boxes = results[0].boxes
                duck_count = len(boxes)
                
                img_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                pil_img = PILImage.fromarray(img_rgb)
                draw = ImageDraw.Draw(pil_img)
                
                for idx, box in enumerate(boxes, start=1):
                    x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                    draw.rectangle([x1, y1, x2, y2], outline=(0, 255, 0), width=3)
                    
                    if box.id is not None:
                        track_id = int(box.id.tolist()[0])
                        text = f"Утка {track_id}"
                    else:
                        text = f"Утка {idx}"
                        
                    try: text_w, text_h = draw.textbbox((0, 0), text, font=font)[2:]
                    except: text_w, text_h = 60, 18
                    draw.rectangle([x1, y1 - text_h - 4, x1 + text_w + 6, y1], fill=(0, 255, 0))
                    draw.text((x1 + 3, y1 - text_h - 2), text, fill=(0, 0, 0), font=font)
                    
                output_frame = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
                out.write(output_frame)
                
                if duck_count >= max_duck_count or best_frame_bgr is None:
                    max_duck_count = duck_count
                    best_frame_bgr = output_frame.copy()
            
            cap.release()
            out.release()
            
            if os.path.exists(temp_input_path):
                os.remove(temp_input_path)
                
            thumb_name = f"thumb_{timestamp_str}_{os.path.splitext(safe_filename)[0]}.jpg"
            thumb_path = os.path.join('static', thumb_name)
            if best_frame_bgr is not None:
                cv2.imwrite(thumb_path, best_frame_bgr)
            else:
                cv2.imwrite(thumb_path, np.zeros((height, width, 3), dtype=np.uint8))
                
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute('INSERT INTO requests (timestamp, duck_count, image_path) VALUES (?, ?, ?)', 
                           (timestamp, max_duck_count, thumb_path))
            
            results_data.append({
                'filename': file.filename,
                'count': max_duck_count,
                'image_url': f"/{thumb_path}",
                'video_url': f"/{video_out_path}",
                'is_video': True
            })
            
    conn.commit()
    conn.close()
    return jsonify(results=results_data)

@app.route('/report', methods=['GET'])
def generate_report():
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute('SELECT timestamp, duck_count, image_path FROM requests ORDER BY timestamp DESC')
    rows = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по уткам"
    
    ws.append(["Дата и время", "Макс. уток / Кол-во", "Путь к скриншоту/фото", "Визуализация"])
    
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 28

    for i, row in enumerate(rows, start=2):
        ws.append([row[0], row[1], row[2]])
        ws.row_dimensions[i].height = 100
        
        img_path = row[2]
        if os.path.exists(img_path):
            try:
                img = ExcelImage(img_path)
                img.width = 180
                img.height = 120
                ws.add_image(img, f'D{i}')
            except Exception as e:
                ws.cell(row=i, column=4, value=f"Ошибка: {e}")
        else:
            ws.cell(row=i, column=4, value="Файл не найден")

    report_path = "static/report.xlsx"
    wb.save(report_path)
    
    return send_file(report_path, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)